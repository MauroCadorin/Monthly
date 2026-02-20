import pandas as pd
import openpyxl
from datetime import datetime
import os
import glob
import warnings

# Suppress openpyxl UserWarning about default style
warnings.filterwarnings("ignore", "Workbook contains no default style", UserWarning)

def clean_bank_operation_and_categorize(operation):
    """
    Clean bank operation name and assign category based on VBA logic for bank statements
    Returns tuple: (cleaned_operation, category, sub_category)
    """
    # Convert to string and handle None values
    if pd.isna(operation) or operation is None:
        return "", "", ""
    
    operation_str = str(operation)
    
    # Clean operation prefixes
    if "BCV-NET " in operation_str:
        operation_str = operation_str.replace("BCV-NET ", "")
    if "VIRT BANC " in operation_str:
        operation_str = operation_str.replace("VIRT BANC ", "")
    if "VIR TWINT " in operation_str:
        operation_str = operation_str.replace("VIR TWINT ", "")
    
    # Bank operation categorization based on VBA logic
    bank_categories = [
        # (search_term, cleaned_name, category, sub_category)
        ("Distalmotion", None, "Salary", ""),
        ("Duol", None, "Chant Cred", "Duol"),
        ("INSTITUT LE CHATELARD", None, "Chant Cred", "Chatelard"),
        ("Leni", None, "Kids Deb", ""),
        ("Assura-Basis", None, "Health", ""),
        ("Etat de Vaud Impôts", None, "Impot", ""),
        ("Swisscom ", None, "Media", ""),
        ("Sunrise", None, "Media", ""),
        ("Salt", None, "Media", ""),
        ("Koloristika", None, "Chant", ""),
        ("Planchamp, Xavier", None, "Rent", ""),
        ("Baptiste Dujardin", None, "Food", ""),
        ("Caisse de pensions de", "Parking", "Car", ""),
        ("PPE LE CAMPUS", None, "Home Crosets", ""),
        ("Romande Energie SA", None, "Home", ""),
        ("Energiapro SA", None, "Home", ""),
        ("PPE SUNDANCE", None, "Home Crosets", ""),
        ("Caisse AVS de la Feder", None, "Alloc", ""),
    ]
    
    # Check each bank operation pattern
    for search_term, cleaned_name, category, sub_category in bank_categories:
        if search_term in operation_str:
            final_name = cleaned_name if cleaned_name else operation_str
            return final_name, category, sub_category
    
    # If no match found, return cleaned operation with empty category
    return operation_str, "", ""

def clean_merchant_and_categorize(merchant):
    """
    Clean merchant name and assign category based on VBA logic
    Returns tuple: (cleaned_merchant, category)
    """
    # Convert to string and handle None values
    if pd.isna(merchant) or merchant is None:
        return "", ""
    
    merchant_str = str(merchant)
    
    # Merchant cleanup and categorization based on VBA logic
    merchant_categories = [
        # (search_term, cleaned_name, category)
        ("Migros", "Migros", "Food"),
        ("Aldi", "Aldi", "Food"),
        ("Denner", "Denner", "Food"),
        ("LAUSANNE10", "LAUSANNE10", "Food"),
        ("Manor", "Manor", "Food"),
        ("Coop", "Coop", "Food"),
        ("Jumbo", "Jumbo", "Home"),
        ("L'Instant Chocolat", "L'Instant Chocolat", "Food"),
        ("APPLE.COM", "APPLE.COM", "Media"),
        ("THE NEW YORK TIMES", "THE NEW YORK TIMES", "Media"),
        ("THE ATHLETIC", "THE ATHLETIC", "Media"),
        ("Tesla", "Tesla", "Car"),
        ("Prime Video", "Prime Video", "Media"),
        ("Sun Store", "Pharmacie-Sunstore", "Health"),
        ("Pharmacie-Sunstore", "Pharmacie-Sunstore", "Health"),
        ("Droguerie Jaquet", "Droguerie Jaquet", "Health"),
        ("Sakura Sushi", "Sakura Sushi", "Food"),
        ("Boutique Ravann", "Boutique Ravann", "Food"),
        ("SBB CFF", "SBB CFF", "Transport"),
        ("Brezelkönig", "Brezelkönig", "Food"),
        ("Zalando", "Zalando", "Clothing"),
        ("BestDrive", "BestDrive", "Car"),
        ("KymeM Cafe", "KymeM Cafe", "Restaurant"),
        ("Pizzeria Vecchia", "Pizzeria Vecchia Napoli", "Restaurant"),
        ("NETFLIX.COM", "NETFLIX.COM", "Media"),
        ("Netflix.com", "NETFLIX.COM", "Media"),
        ("Salt", "Salt Mobile SA", "Media"),
        ("salt.ch", "Salt Mobile SA", "Media"),
        ("Patreon", "Patreon", "Media"),
        ("Association Golf de La Puidoux", "Association Golf de La Puidoux", "Hobby"),
        ("Exotic Food Center", "Exotic Food Center", "Food"),
        ("Aux Merveilleux", "Aux Merveilleux", "Food"),
        ("Appunto Rest.", "Appunto Rest.", "Food"),
        ("QoQa Services SA", "QoQa", "?"),
        ("DAZN", "DAZN", "Media"),
        ("URUMQI", "URUMQI", "Food"),
        ("La Cavagne", "La Cavagne", "Food"),
    ]
    
    # Check each merchant pattern
    for search_term, cleaned_name, category in merchant_categories:
        if search_term in merchant_str:
            return cleaned_name, category
    
    # If no match found, return original merchant with empty category
    return merchant_str, ""

def parse_date(date_str):
    """
    Parse date from DD-MM-YYYY format to datetime object
    """
    try:
        # Try different date formats
        if isinstance(date_str, datetime):
            return date_str
        
        date_str = str(date_str).strip()
        
        # Try DD-MM-YYYY format (from CSV)
        try:
            return datetime.strptime(date_str, "%d-%m-%Y")
        except:
            pass
            
        # Try DD.MM.YYYY format (from Excel)
        try:
            return datetime.strptime(date_str, "%d.%m.%Y")
        except:
            pass
            
        # Try other common formats
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"]:
            try:
                return datetime.strptime(date_str, fmt)
            except:
                continue
                
        return None
    except:
        return None

def process_account_statements():
    """
    Process account statements from CSV files starting with 'account-statement_'
    """
    # Find all account statement CSV files
    account_files = glob.glob("account-statement_*.csv")
    
    if not account_files:
        print("No account statement CSV files found matching pattern 'account-statement_*.csv'")
        return
    
    print(f"Found {len(account_files)} account statement files: {account_files}")
    
    # Use hardcoded Excel file
    excel_file = "LISTE DES OPÉRATIONS-2025.xlsm"
    
    if not os.path.exists(excel_file):
        print(f"Excel file {excel_file} not found!")
        return
    
    print(f"Using Excel file: {excel_file}")
    
    # Process all account statement files
    all_transactions = []
    
    for csv_file in account_files:
        try:
            print(f"Processing {csv_file}...")
            df = pd.read_csv(csv_file)
            
            # Filter for "Card Payment" rows only
            card_payments = df[df['Type'] == 'Card Payment']
            
            for _, row in card_payments.iterrows():
                try:
                    # Get completed date from column D (index 3)
                    completed_date_str = row['Completed Date']
                    if pd.isna(completed_date_str):
                        continue
                    
                    # Parse date and convert to dd.mm.yyyy format (remove time)
                    # Expected format: "2025-11-13 14:07:59"
                    completed_date = datetime.strptime(completed_date_str, "%Y-%m-%d %H:%M:%S")
                    
                    # Get merchant from column E (Description)
                    merchant = row['Description']
                    if pd.isna(merchant):
                        continue
                    
                    # Get amount from column F and make it positive
                    amount = row['Amount']
                    if pd.isna(amount):
                        continue
                    
                    amount = abs(float(amount))  # Make positive
                    
                    # Clean merchant and get category
                    cleaned_merchant, category = clean_merchant_and_categorize(merchant)
                    
                    all_transactions.append({
                        'date': completed_date,
                        'merchant': cleaned_merchant,
                        'amount': amount,
                        'category': category
                    })
                    
                except Exception as e:
                    print(f"Error processing row in {csv_file}: {e}")
                    continue
        
        except Exception as e:
            print(f"Error reading {csv_file}: {e}")
            continue
    
    print(f"Processed {len(all_transactions)} card payment transactions from account statements")
    
    if not all_transactions:
        print("No card payment transactions found to process")
        return
    
    # Sort transactions from newest to oldest
    all_transactions.sort(key=lambda x: x['date'], reverse=True)
    
    try:
        # Load Excel file
        print(f"Loading {excel_file}...")
        workbook = openpyxl.load_workbook(excel_file, keep_vba=True)
        
        # Create or get Revolut worksheet
        revolut_sheet_name = "Revolut"
        if revolut_sheet_name in workbook.sheetnames:
            worksheet = workbook[revolut_sheet_name]
        else:
            worksheet = workbook.create_sheet(revolut_sheet_name)
            # Add headers to Revolut sheet
            worksheet['A1'] = "Date"
            worksheet['B1'] = "Merchant"
            worksheet['C1'] = "Amount"
            worksheet['D1'] = "Category"
            worksheet['E1'] = "Reason"
        
        # Create or get duplicates worksheet
        duplicate_sheet_name = "Duplicates"
        if duplicate_sheet_name in workbook.sheetnames:
            duplicate_worksheet = workbook[duplicate_sheet_name]
        else:
            duplicate_worksheet = workbook.create_sheet(duplicate_sheet_name)
            # Add headers to duplicate sheet
            duplicate_worksheet['A1'] = "Date"
            duplicate_worksheet['B1'] = "Merchant"
            duplicate_worksheet['C1'] = "Amount"
            duplicate_worksheet['D1'] = "Category"
            duplicate_worksheet['E1'] = "Reason"
        
        # Find the last row with data in column A
        last_row = 1
        for row in range(1, worksheet.max_row + 1):
            if worksheet[f'A{row}'].value is not None:
                last_row = row
        
        print(f"Found {last_row} existing rows in Excel")
        
        # Get existing data for duplicate check
        existing_data = set()
        for row in range(1, last_row + 1):
            date_val = worksheet[f'A{row}'].value
            merchant_val = worksheet[f'B{row}'].value
            amount_val = worksheet[f'C{row}'].value
            
            if date_val and merchant_val and amount_val:
                # Skip header row or non-numeric amounts
                try:
                    amount_float = float(amount_val)
                except (ValueError, TypeError):
                    continue
                
                # Convert date to string for comparison
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)
                
                existing_data.add((date_str, str(merchant_val), amount_float))
        
        # Find last row in duplicate sheet
        duplicate_last_row = 1
        for row in range(1, duplicate_worksheet.max_row + 1):
            if duplicate_worksheet[f'A{row}'].value is not None:
                duplicate_last_row = row
        
        # Add new transactions (avoiding duplicates)
        new_transactions_added = 0
        duplicates_found = 0
        next_row = last_row + 1
        next_duplicate_row = duplicate_last_row + 1
        
        # Find oldest transaction date among new transactions to add
        new_transactions_to_add = []
        for transaction in all_transactions:
            trans_tuple = (
                transaction['date'].strftime("%Y-%m-%d"),
                transaction['merchant'],
                float(transaction['amount'])
            )
            if trans_tuple not in existing_data:
                new_transactions_to_add.append(transaction)
        
        oldest_date = min([t['date'] for t in new_transactions_to_add]) if new_transactions_to_add else None
        
        for transaction in all_transactions:
            # Create tuple for duplicate check
            trans_tuple = (
                transaction['date'].strftime("%Y-%m-%d"),
                transaction['merchant'],
                float(transaction['amount'])
            )
            
            # Check if duplicate
            if trans_tuple in existing_data:
                # Add to duplicate sheet
                date_str = transaction['date'].strftime("%d.%m.%Y")
                duplicate_worksheet[f'A{next_duplicate_row}'] = date_str
                duplicate_worksheet[f'A{next_duplicate_row}'].number_format = '@'  # Force text format
                
                duplicate_worksheet[f'B{next_duplicate_row}'] = transaction['merchant']
                duplicate_worksheet[f'C{next_duplicate_row}'] = transaction['amount']
                duplicate_worksheet[f'D{next_duplicate_row}'] = transaction['category']
                duplicate_worksheet[f'E{next_duplicate_row}'] = "Duplicate from account statement"
                duplicates_found += 1
                next_duplicate_row += 1
                continue
            
            # Add to main Excel sheet with dd.mm.yyyy format
            date_str = transaction['date'].strftime("%d.%m.%Y")
            worksheet[f'A{next_row}'] = date_str
            worksheet[f'A{next_row}'].number_format = '@'  # Force text format
            
            worksheet[f'B{next_row}'] = transaction['merchant']
            worksheet[f'C{next_row}'] = transaction['amount']
            
            # Add category in column E if it exists
            if transaction['category']:
                worksheet[f'E{next_row}'] = transaction['category']
            
            # Make oldest transactions bold
            if oldest_date and transaction['date'] == oldest_date:
                from openpyxl.styles import Font
                bold_font = Font(bold=True)
                worksheet[f'A{next_row}'].font = bold_font
                worksheet[f'B{next_row}'].font = bold_font
                worksheet[f'C{next_row}'].font = bold_font
                if transaction['category']:
                    worksheet[f'E{next_row}'].font = bold_font
            
            existing_data.add(trans_tuple)
            new_transactions_added += 1
            next_row += 1
        
        # Save the workbook
        print(f"Adding {new_transactions_added} new account statement transactions to Revolut sheet...")
        workbook.save(excel_file)
        workbook.close()
        
        print(f"Account statement processing completed!")
        print(f"Added {new_transactions_added} new transactions to Revolut sheet.")
        print(f"Found {duplicates_found} duplicates (added to 'Duplicates' sheet).")
        
    except Exception as e:
        print(f"Error processing account statements: {str(e)}")
        import traceback
        traceback.print_exc()

def process_bank_statements():
    """
    Process bank statements from LISTE DES OPÉRATIONS files
    """
    # Find input file with bracket pattern [dd-mm-yyyy]
    input_files = glob.glob("LISTE DES OPÉRATIONS *.xlsx")
    
    # Filter files that actually contain brackets in the name
    bracket_files = [f for f in input_files if '[' in f and ']' in f]
    
    if not bracket_files:
        print("No bank statement files found matching pattern 'LISTE DES OPÉRATIONS [*].xlsx'")
        return
    
    # Use the first matching file as input
    input_file = bracket_files[0]
    print(f"Using input file: {input_file}")
    
    # Use hardcoded output file
    output_file = "LISTE DES OPÉRATIONS-2025.xlsm"
    
    if not os.path.exists(output_file):
        print(f"Output file {output_file} not found!")
        return
    
    print(f"Using output file: {output_file}")
    
    try:
        print(f"Processing {input_file}...")
        
        # Read bank statement input file with correct header row
        df = pd.read_excel(input_file, header=8)  # Row 9 contains data

        # Load output Excel file
        print(f"Loading {output_file}...")
        workbook = openpyxl.load_workbook(output_file, keep_vba=True)
        
        # Check 2025 worksheet to get the latest transaction date
        latest_date = None
        sheet_2025_name = "2025"
        if sheet_2025_name in workbook.sheetnames:
            sheet_2025 = workbook[sheet_2025_name]
            # Find the latest date in column A of 2025 worksheet
            for row in range(2, sheet_2025.max_row + 1):  # Skip header row
                date_val = sheet_2025[f'A{row}'].value
                if date_val:
                    try:
                        if isinstance(date_val, datetime):
                            current_date = date_val
                        else:
                            current_date = parse_date(str(date_val))
                        
                        if current_date and (latest_date is None or current_date > latest_date):
                            latest_date = current_date
                    except:
                        continue
        
        if latest_date:
            print(f"Found latest date in 2025 worksheet: {latest_date.strftime('%d.%m.%Y')}")
            print(f"Will only process transactions after this date")
        else:
            print("No existing dates found in 2025 worksheet, will process all transactions")
          
        # Process bank transactions
        bank_transactions = []
        for index, row in df.iterrows():
                try:
                    # Parse date - get from column A (index 0)
                    execution_date = row.iloc[0]  # Column A
                    if pd.isna(execution_date):
                        continue
                        
                    # Use the parse_date function for consistent date parsing
                    execution_date = parse_date(execution_date)
                    if execution_date is None:
                        continue
                    
                    operation = row.iloc[1]  # Column B - Opérations
                    if pd.isna(operation):
                        continue
                        
                    # Get debit from column C and credit from column D as separate values
                    debit = row.iloc[2] if pd.notna(row.iloc[2]) else 0  # Column C - Débit
                    credit = row.iloc[3] if pd.notna(row.iloc[3]) else 0  # Column D - Crédit
                    
                    # Skip if both debit and credit are zero
                    if debit == 0 and credit == 0:
                        continue
                    
                    # Clean operation and get category
                    cleaned_operation, category, sub_category = clean_bank_operation_and_categorize(operation)
                    
                    # Only add transaction if it's after the latest date from 2025 worksheet
                    if latest_date is None or execution_date > latest_date:
                        bank_transactions.append({
                            'date': execution_date,
                            'operation': cleaned_operation,
                            'debit': float(debit) if debit != 0 else 0,
                            'credit': float(credit) if credit != 0 else 0,
                            'category': category,
                            'sub_category': sub_category
                        })
                    else:
                        print(f"Skipping transaction from {execution_date.strftime('%d.%m.%Y')} (before latest date)")
                    
                except Exception as e:
                    print(f"Error processing row {index}: {e}")
                    continue
        
        print(f"Processed {len(bank_transactions)} bank transactions from input file")
        
        if not bank_transactions:
            print("No bank transactions found to process")
            return
        
        # Sort transactions from newest to oldest
        bank_transactions.sort(key=lambda x: x['date'], reverse=True)

        # Get or create the BCV worksheet
        sheet_bcv_name = "BCV"
        if sheet_bcv_name in workbook.sheetnames:
            sheet_bcv = workbook[sheet_bcv_name]
        else:
            sheet_bcv = workbook.create_sheet(sheet_bcv_name)
            # Add headers to BCV sheet
            sheet_bcv['A1'] = "Date"
            sheet_bcv['B1'] = "Operation"
            sheet_bcv['C1'] = "Debit"
            sheet_bcv['D1'] = "Credit"
            sheet_bcv['E1'] = "Category"
            sheet_bcv['F1'] = "Sub Category"
        
        # Get existing data from BCV tab for duplicate check
        existing_data = set()
        for row in range(1, sheet_bcv.max_row + 1):
            date_val = sheet_bcv[f'A{row}'].value
            operation_val = sheet_bcv[f'B{row}'].value
            debit_val = sheet_bcv[f'C{row}'].value
            credit_val = sheet_bcv[f'D{row}'].value
            
            if date_val and operation_val and (debit_val or credit_val):
                # Skip header row or non-numeric amounts
                try:
                    debit_float = float(debit_val) if debit_val else 0
                    credit_float = float(credit_val) if credit_val else 0
                except (ValueError, TypeError):
                    continue
                
                # Convert date to string for comparison
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)
                
                existing_data.add((date_str, str(operation_val), debit_float, credit_float))
        
        # Find the last row with data in BCV tab
        last_row_bcv = 1
        for row in range(1, sheet_bcv.max_row + 1):
            if sheet_bcv[f'A{row}'].value is not None:
                last_row_bcv = row
        
        # Add new transactions (avoiding duplicates)
        new_transactions_added = 0
        duplicates_skipped = 0
        next_row = last_row_bcv + 1
        
        # Find oldest transaction date among new transactions to add
        new_bank_transactions_to_add = []
        for transaction in bank_transactions:
            trans_tuple = (
                transaction['date'].strftime("%Y-%m-%d"),
                transaction['operation'],
                float(transaction['debit']),
                float(transaction['credit'])
            )
            if trans_tuple not in existing_data:
                new_bank_transactions_to_add.append(transaction)
        
        oldest_bank_date = min([t['date'] for t in new_bank_transactions_to_add]) if new_bank_transactions_to_add else None
        
        for transaction in bank_transactions:
            # Create tuple for duplicate check
            trans_tuple = (
                transaction['date'].strftime("%Y-%m-%d"),
                transaction['operation'],
                float(transaction['debit']),
                float(transaction['credit'])
            )
            
            # Check if duplicate
            if trans_tuple in existing_data:
                duplicates_skipped += 1
                continue
            
            # Add to BCV sheet with dd.mm.yyyy format
            date_str = transaction['date'].strftime("%d.%m.%Y")
            sheet_bcv[f'A{next_row}'] = date_str
            sheet_bcv[f'A{next_row}'].number_format = '@'  # Force text format
            
            sheet_bcv[f'B{next_row}'] = transaction['operation']
            sheet_bcv[f'C{next_row}'] = transaction['debit'] if transaction['debit'] != 0 else ""
            sheet_bcv[f'D{next_row}'] = transaction['credit'] if transaction['credit'] != 0 else ""
            
            if transaction['category']:
                sheet_bcv[f'E{next_row}'] = transaction['category']
            
            if transaction['sub_category']:
                sheet_bcv[f'F{next_row}'] = transaction['sub_category']
            
            # Make oldest transactions bold
            if oldest_bank_date and transaction['date'] == oldest_bank_date:
                from openpyxl.styles import Font
                bold_font = Font(bold=True)
                sheet_bcv[f'A{next_row}'].font = bold_font
                sheet_bcv[f'B{next_row}'].font = bold_font
                sheet_bcv[f'C{next_row}'].font = bold_font
                sheet_bcv[f'D{next_row}'].font = bold_font
                sheet_bcv[f'E{next_row}'].font = bold_font
                sheet_bcv[f'F{next_row}'].font = bold_font
            
            existing_data.add(trans_tuple)
            new_transactions_added += 1
            next_row += 1
        
        # Save the workbook
        print(f"Adding {new_transactions_added} new bank transactions to '{sheet_bcv_name}' sheet...")
        workbook.save(output_file)
        workbook.close()
        
        print(f"Bank statement processing completed!")
        print(f"Added {new_transactions_added} new transactions to BCV tab.")
        print(f"Skipped {duplicates_skipped} duplicates.")
        
    except Exception as e:
        print(f"Error processing bank statements: {str(e)}")
        import traceback
        traceback.print_exc()

def process_transactions():
    """
    Main function to process transactions from CSV and update Excel file
    """
    # File paths
    csv_file = "transactions.csv"
    
    # Use hardcoded Excel file
    excel_file = "LISTE DES OPÉRATIONS-2025.xlsm"
    
    if not os.path.exists(excel_file):
        print(f"Excel file {excel_file} not found!")
        return
    
    print(f"Using Excel file: {excel_file}")
    
    # Check if files exist
    if not os.path.exists(csv_file):
        print(f"Error: {csv_file} not found!")
        return
    
    try:
        # Read CSV file
        print("Reading transactions.csv...")
        df = pd.read_csv(csv_file)
        
        # Extract required columns
        transactions = []
        for _, row in df.iterrows():
            # Parse booking date using the improved parse_date function
            booking_date = parse_date(row['Booking date'])
            if booking_date is None:
                print(f"Warning: Could not parse date '{row['Booking date']}'")
                continue
            
            merchant = row['Merchant']
            transaction_type = row['Type']
            
            # Parse amount, handling potential string formatting
            try:
                amount = float(row['Amount (CHF)'])
                # If type is Credit, make amount negative
                if 'Credit' in str(transaction_type):
                    amount = -amount
            except (ValueError, TypeError):
                continue
            
            # Clean merchant and get category
            cleaned_merchant, category = clean_merchant_and_categorize(merchant)
            
            transactions.append({
                'date': booking_date,
                'merchant': cleaned_merchant,
                'amount': amount,
                'category': category
            })
        
        print(f"Processed {len(transactions)} transactions from CSV")
        
        # Sort transactions from newest to oldest
        transactions.sort(key=lambda x: x['date'], reverse=True)
        
        # Load Excel file
        print(f"Loading {excel_file}...")
        workbook = openpyxl.load_workbook(excel_file, keep_vba=True)
        
        # Create or get Carte Cred worksheet
        carte_cred_sheet_name = "Carte Cred"
        if carte_cred_sheet_name in workbook.sheetnames:
            worksheet = workbook[carte_cred_sheet_name]
        else:
            worksheet = workbook.create_sheet(carte_cred_sheet_name)
            # Add headers to Carte Cred sheet
            worksheet['A1'] = "Date"
            worksheet['B1'] = "Merchant"
            worksheet['C1'] = "Amount"
            worksheet['D1'] = "Empty"
            worksheet['E1'] = "Category"
        
        # Create or get duplicates worksheet
        duplicate_sheet_name = "Duplicates"
        if duplicate_sheet_name in workbook.sheetnames:
            duplicate_worksheet = workbook[duplicate_sheet_name]
        else:
            duplicate_worksheet = workbook.create_sheet(duplicate_sheet_name)
            # Add headers to duplicate sheet
            duplicate_worksheet['A1'] = "Date"
            duplicate_worksheet['B1'] = "Merchant"
            duplicate_worksheet['C1'] = "Amount"
            duplicate_worksheet['D1'] = "Empty"
            duplicate_worksheet['E1'] = "Category"
        
        # Find the last row with data in column A
        last_row = 1
        for row in range(1, worksheet.max_row + 1):
            if worksheet[f'A{row}'].value is not None:
                last_row = row
        
        print(f"Found {last_row} existing rows in Excel")
        
        # Get existing data for duplicate check
        existing_data = set()
        for row in range(1, last_row + 1):
            date_val = worksheet[f'A{row}'].value
            merchant_val = worksheet[f'B{row}'].value
            amount_val = worksheet[f'C{row}'].value
            
            if date_val and merchant_val and amount_val:
                # Skip header row or non-numeric amounts
                try:
                    amount_float = float(amount_val)
                except (ValueError, TypeError):
                    continue
                
                # Convert date to string for comparison
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)
                
                existing_data.add((date_str, str(merchant_val), amount_float))
        
        # Find last row in duplicate sheet
        duplicate_last_row = 1
        for row in range(1, duplicate_worksheet.max_row + 1):
            if duplicate_worksheet[f'A{row}'].value is not None:
                duplicate_last_row = row
        
        # Add new transactions (avoiding duplicates)
        new_transactions_added = 0
        duplicates_found = 0
        next_row = last_row + 1
        next_duplicate_row = duplicate_last_row + 1
        
        # Find oldest transaction date among new transactions to add
        new_cc_transactions_to_add = []
        for transaction in transactions:
            trans_tuple = (
                transaction['date'].strftime("%Y-%m-%d"),
                transaction['merchant'],
                float(transaction['amount'])
            )
            if trans_tuple not in existing_data:
                new_cc_transactions_to_add.append(transaction)
        
        oldest_cc_date = min([t['date'] for t in new_cc_transactions_to_add]) if new_cc_transactions_to_add else None
        
        for transaction in transactions:
            # Create tuple for duplicate check
            trans_tuple = (
                transaction['date'].strftime("%Y-%m-%d"),
                transaction['merchant'],
                float(transaction['amount'])
            )
            
            # Check if duplicate
            if trans_tuple in existing_data:
                # Add to duplicate sheet
                date_str = transaction['date'].strftime("%d.%m.%Y")
                duplicate_worksheet[f'A{next_duplicate_row}'] = date_str
                duplicate_worksheet[f'A{next_duplicate_row}'].number_format = '@'  # Force text format
                
                duplicate_worksheet[f'B{next_duplicate_row}'] = transaction['merchant']
                duplicate_worksheet[f'C{next_duplicate_row}'] = transaction['amount']
                duplicate_worksheet[f'D{next_duplicate_row}'] = transaction['category']
                duplicate_worksheet[f'E{next_duplicate_row}'] = "Duplicate entry"
                duplicates_found += 1
                next_duplicate_row += 1
                continue
            
            # Add to main Excel sheet with dd.mm.yyyy format
            date_str = transaction['date'].strftime("%d.%m.%Y")
            worksheet[f'A{next_row}'] = date_str
            worksheet[f'A{next_row}'].number_format = '@'  # Force text format
            
            worksheet[f'B{next_row}'] = transaction['merchant']
            worksheet[f'C{next_row}'] = transaction['amount']
            
            # Add category in column E if it exists
            if transaction['category']:
                worksheet[f'E{next_row}'] = transaction['category']
            
            # Make oldest transactions bold
            if oldest_cc_date and transaction['date'] == oldest_cc_date:
                from openpyxl.styles import Font
                bold_font = Font(bold=True)
                worksheet[f'A{next_row}'].font = bold_font
                worksheet[f'B{next_row}'].font = bold_font
                worksheet[f'C{next_row}'].font = bold_font
                if transaction['category']:
                    worksheet[f'E{next_row}'].font = bold_font
            
            existing_data.add(trans_tuple)
            new_transactions_added += 1
            next_row += 1
        
        # Save the workbook
        print(f"Adding {new_transactions_added} new transactions to Carte Cred sheet...")
        workbook.save(excel_file)
        workbook.close()
        
        print(f"Successfully processed! Added {new_transactions_added} new transactions to Carte Cred sheet.")
        print(f"Found {duplicates_found} duplicates (added to 'Duplicates' sheet).")
        print(f"Total rows in Carte Cred sheet: {next_row - 1}")
        print(f"Total rows in duplicates sheet: {next_duplicate_row - 1}")
        
    except Exception as e:
        print(f"Error processing transactions: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    print("\nProcessing bank statements...")
    process_bank_statements()

    print("Processing credit card transactions...")
    process_transactions()
        
    print("\nProcessing account statements...")
    process_account_statements()
